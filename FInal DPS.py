# ==============================
# Import Required Packages
# ==============================
import openai  # OpenAI API
import time    # For delays
import os      # File operations
import shutil  # File management
from selenium import webdriver  # Selenium for browser automation
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.common.exceptions import TimeoutException, WebDriverException
from selenium.webdriver.common.action_chains import ActionChains  # For right-click operations
import openpyxl  # Excel sheet management
from PIL import Image  # Image processing
from pydrive.auth import GoogleAuth  # Google Drive upload
from pydrive.drive import GoogleDrive
import logging  # For logging process steps
import pathlib  # Path operations

import logging  # For logging process steps
import pathlib  # Path operations

# ==============================
# Parameters (User Inputs)
# ==============================

# --- Selenium Settings ---
DISCORD_LOGIN_EMAIL = "benchekroune.yassine7@gmail.com"  # Replace with Discord email
DISCORD_LOGIN_PASSWORD = "Malzahar77."  # Replace with Discord password
MIDJOURNEY_WEB_URL = "https://www.midjourney.com"  # MidJourney Web URL

# --- Edge Driver Path ---
EDGE_DRIVER_PATH = r"C:\WebDriver\msedgedriver.exe"  # Replace with actual EdgeDriver path
DEBUG_PORT = "9222"  # Example debugging port

# --- Output Folders ---
BASE_OUTPUT_FOLDER = r"C:\Users\ULTRAPC\Desktop\Digital Paper Store\Digital Paper Store - Raw Folders"  # Base directory for product folders
DOWNLOADS_FOLDER = os.path.expanduser("~\\Downloads")  # Downloads folder
SEAMLESS_PATTERN_FOLDER = r"C:\Users\ULTRAPC\Desktop\Digital Paper Store\Digital Paper Store - Seamless Pattern"
DIGITAL_PAPER_FOLDER = r"C:\Users\ULTRAPC\Desktop\Digital Paper Store\Digital Paper Store - Digital Paper"

# --- Excel File ---
INPUT_EXCEL_FILE = r"C:\Users\ULTRAPC\Downloads\template (4).xlsx"  # Replace with your Excel input file

# --- General Script Settings ---
NUMBER_OF_PROMPTS_PER_PRODUCT = 10  # Number of prompts to send for each product
WAIT_TIME_BETWEEN_PROMPTS = 3  # Time (seconds) to wait between prompts in Selenium
WAIT_BETWEEN_IMAGE_CHECKS = 45  # Time to wait before checking for images
MAX_IMAGE_WAIT_TIME = 900  # Maximum wait time for image generation (seconds)

# URLs for the "Organize" and "Explore" pages
ORGANIZE_PAGE_URL = "https://www.midjourney.com/archive"
EXPLORE_PAGE_URL = "https://www.midjourney.com/explore?tab=top"

# Add this function to navigate directly to the "Organize" page using the URL
def navigate_to_organize_page(driver):
    try:
        driver.get(ORGANIZE_PAGE_URL)
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.TAG_NAME, "textarea")))
        print("Navigated to the Organize page.")
    except Exception as e:
        logging.error(f"Error navigating to the Organize page: {e}")
        raise

# ==============================
# Read Prompts and Details from Excel
# ==============================
def read_prompts_from_excel(file_path):
    """
    Read prompts, Product Type, Category, and Theme from the specified Excel file.
    Ensures 'Prompts' column is accurately targeted by name.
    """
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]  # Read first row to identify headers

        prompts_index = headers.index("Prompts")
        theme_index = headers.index("Theme")
        category_index = headers.index("Category")
        product_type_index = headers.index("Product Type")

        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[prompts_index]:
                prompts = [line.strip() for line in row[prompts_index].splitlines() if line.strip()]
                data.append({
                    "Prompts": prompts,
                    "Theme": row[theme_index],
                    "Category": row[category_index],
                    "Product Type": row[product_type_index],
                    "Row Index": row[0]  # Assuming the first column is the row index
                })
        return data
    except Exception as e:
        logging.error(f"Error reading Excel file: {e}")
        return []

# ==============================
# Edge Browser Debugging Setup
# ==============================
def launch_edge_debug():
    print("Launching Edge in debugging mode. Log in to MidJourney manually.")
    os.system(f'start msedge --remote-debugging-port={DEBUG_PORT} --user-data-dir="C:\\EdgeProfile"')

def connect_to_existing_edge():
    try:
        options = EdgeOptions()
        options.use_chromium = True
        options.debugger_address = f"127.0.0.1:{DEBUG_PORT}"
        driver_service = EdgeService(EDGE_DRIVER_PATH)
        return webdriver.Edge(service=driver_service, options=options)
    except Exception as e:
        print("No Edge window detected. Launching Edge in debugging mode...")
        launch_edge_debug()
        time.sleep(10)
        return connect_to_existing_edge()

# ==============================
# MidJourney Automation with Selenium
# ==============================
# Add this function to verify and navigate to the "Organize" page
def ensure_on_organize_page(driver):
    try:
        ORGANIZE_PAGE_URL = "https://www.midjourney.com/archive"
        if "organize" not in driver.current_url.lower():
            print("Navigating to the Organize page...")
            driver.get(ORGANIZE_PAGE_URL)
            WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.TAG_NAME, "textarea")))
            print("Successfully navigated to the Organize page.")
    except Exception as e:
        logging.error(f"Error navigating to the Organize page: {e}")
        raise

# Add this function to check if the current page is a verification page
def is_verification_page(driver):
    try:
        # Check for specific elements that are unique to the verification page
        verification_elements = [
            (By.XPATH, "//p[@id='TBuuD2' and contains(text(), 'Verify you are human by completing the action below.')]"),
            (By.XPATH, "//span[contains(@class, 'cb-lb-t') and contains(text(), 'Verify you are human')]"),
            (By.XPATH, "//iframe[contains(@src, 'captcha')]"),
            (By.XPATH, "//div[contains(@class, 'verification')]"),
            (By.XPATH, "//div[contains(text(), 'Verify you are human')]"),
            (By.XPATH, "//div[contains(text(), 'human verification')]")
        ]
        for by, value in verification_elements:
            if driver.find_elements(by, value):
                logging.info("Verification page detected.")
                return True

        logging.info("Not a verification page.")
        return False
    except Exception as e:
        logging.error(f"Error checking for verification page: {e}")
        return False

# Update the send_prompts_to_midjourney function
def send_prompts_to_midjourney(driver, data):
    try:
        driver = connect_to_existing_edge()
        driver.get(MIDJOURNEY_WEB_URL)

        # Ensure we are on the Organize page
        ensure_on_organize_page(driver)

        # Refresh the page to check for verification before downloading images
        driver.refresh()
        time.sleep(5)  # Wait for the page to load
        if is_verification_page(driver):
                input("Please complete the verification and press Enter to continue...")
        

        for idx, entry in enumerate(data):
            prompts = entry.get('Prompts', [])
            # Extract Theme, Category, and Product Type from the Excel entry
            theme = entry.get("Theme", "").strip()
            category = entry.get("Category", "").strip()
            product_type = entry.get("Product Type", "").strip()
            
            # Construct the product_name and raw_folder_name dynamically
            product_name = f"{theme} - {category} - {product_type}"
            raw_folder_name = f"Raw {product_name}"
            
            # Sanitize names to ensure they are file/folder safe
            sanitized_product_name = sanitize_name(product_name)
            sanitized_raw_folder_name = sanitize_name(raw_folder_name)
            
            # Debugging: Print the names for verification
            print(f"Sending prompts for: {sanitized_product_name}")
            print(f"Raw folder name: {sanitized_raw_folder_name}")

            # Refresh the page to check for verification
            driver.refresh()
            time.sleep(5)  # Wait for the page to load

            # Check if redirected to verification page
            if is_verification_page(driver):
                input("Please complete the verification and press Enter to continue...")

            for prompt_idx, prompt in enumerate(prompts):
                print(f"Submitting Prompt {prompt_idx+1}: {prompt}")
                try:
                    text_area = driver.find_element(By.TAG_NAME, "textarea")
                    text_area.clear()
                    text_area.send_keys(prompt)
                    text_area.send_keys(Keys.RETURN)
                    time.sleep(WAIT_TIME_BETWEEN_PROMPTS)
                except Exception as e:
                    logging.error(f"Error submitting prompt: {e}")
                    continue

            wait_for_last_image_to_generate(driver)

            # Refresh the page to check for verification before downloading images
            driver.refresh()
            time.sleep(5)  # Wait for the page to load
            if is_verification_page(driver):
                input("Please complete the verification and press Enter to continue...")

            

            raw_folder_path = download_images(driver, sanitized_raw_folder_name)
            processed_folder_path = os.path.join(
                SEAMLESS_PATTERN_FOLDER if entry['Product Type'] == "Seamless Pattern" else DIGITAL_PAPER_FOLDER,
                f"{sanitized_product_name}-DPS"
            )
            process_images(raw_folder_path, processed_folder_path, entry['Product Type'])
            print(f"Folder successfully processed: {processed_folder_path}")

            # Refresh the page after downloading images
            driver.refresh()
            time.sleep(5)  # Wait for the page to load
            if is_verification_page(driver):
                input("Please complete the verification and press Enter to continue...")
    except Exception as e:
        logging.error(f"Error during prompt submission: {e}")

# Add this function to verify and navigate to the "Organize" page
def ensure_on_organize_page(driver):
    try:
        ORGANIZE_PAGE_URL = "https://www.midjourney.com/archive"
        if "archive" not in driver.current_url.lower():
            print("Navigating to the Organize page...")
            driver.get(ORGANIZE_PAGE_URL)
            WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.TAG_NAME, "textarea")))
            # Refresh the page to check for verification before downloading images
            driver.refresh()
            time.sleep(5)  # Wait for the page to load
            
        if is_verification_page(driver):
                input("Please complete the verification and press Enter to continue...")
        print("Successfully navigated to the Organize page.")

    except Exception as e:
        logging.error(f"Error navigating to the Organize page: {e}")
        raise

# ==============================
# Wait and Process Images
def wait_for_last_image_to_generate(driver):
    """
    Wait for a fixed timeout to ensure all images have time to generate.
    Then proceed to select and download the last 40 images.
    """
    fixed_wait_time = 1 * 60  # 14 minutes timeout in seconds
    print(f"⏳ Waiting for {fixed_wait_time // 60} minutes to allow images to generate...")

    # Wait for the fixed timeout
    time.sleep(fixed_wait_time)

    print("✅ Timeout reached. Assuming all images are generated.")

def sanitize_name(name):
    return "".join(c for c in name if c.isalnum() or c in (" ", "-", "_")).strip()

def download_images(driver, raw_folder_name):
    print("Selecting and downloading images...")
    images = driver.find_elements(By.CSS_SELECTOR, "img")[:4]  # Top 40 images
    for image in images:
        ActionChains(driver).key_down(Keys.SHIFT).click(image).key_up(Keys.SHIFT).perform()

    download_button = driver.find_element(By.XPATH, "//button[contains(text(), 'Download')]")
    download_button.click()
    time.sleep(45)  # Wait for download
    if download_button is None:
        for _ in range(3):  # Retry up to 3 times
            try:
                download_button = WebDriverWait(driver, 30).until(
                    EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'Download')]"))
                )
                break
            except Exception as e:
                logging.error(f"Error locating download button: {e}")
                time.sleep(5)  # Wait before retrying

        if download_button:
            download_button.click()
            time.sleep(45)  # Wait for download
        else:
            logging.error("Failed to locate the download button after multiple attempts.")
            # Close the download section dropdown if it is open
            try:
                close_button = driver.find_element(By.XPATH, "//button[contains(@aria-label, 'Close')]")
                close_button.click()
                time.sleep(2)  # Wait for the dropdown to close
            except Exception as e:
                logging.error(f"Error closing the download dropdown: {e}")

            # Refresh the page and reselect images
            driver.refresh()
            time.sleep(5)  # Wait for the page to load
            if is_verification_page(driver):
                input("Please complete the verification and press Enter to continue...")

            # Reselect images
            images = driver.find_elements(By.CSS_SELECTOR, "img")[:40]  # Top 40 images
            for image in images:
                ActionChains(driver).key_down(Keys.SHIFT).click(image).key_up(Keys.SHIFT).perform()

            # Retry locating the download button
            try:
                download_button = WebDriverWait(driver, 30).until(
                    EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'Download')]"))
                )
                download_button.click()
                time.sleep(45)  # Wait for download
            except Exception as e:
                logging.error(f"Error locating download button after retry: {e}")

    # Extract and rename folder
    zip_file_path = max(pathlib.Path(DOWNLOADS_FOLDER).glob("*.zip"), key=os.path.getctime)
    extracted_folder_path = os.path.join(BASE_OUTPUT_FOLDER, raw_folder_name)
    if not os.path.exists(extracted_folder_path):
        os.makedirs(extracted_folder_path)
    shutil.unpack_archive(zip_file_path, extracted_folder_path)
    os.remove(zip_file_path)  # Cleanup zip file

    return extracted_folder_path

def process_images(raw_folder_path, processed_folder_path, product_type):
    # Define processed folder destination
    destination_folder = SEAMLESS_PATTERN_FOLDER if product_type == "Seamless Pattern" else DIGITAL_PAPER_FOLDER
    if not os.path.exists(processed_folder_path):
        os.makedirs(processed_folder_path)
    for idx, filename in enumerate(os.listdir(raw_folder_path)):
        if filename.lower().endswith('.png'):
            input_path = os.path.join(raw_folder_path, filename)
            output_path = os.path.join(processed_folder_path, filename)
            with Image.open(input_path) as img:
                img = img.resize((3600, 3600), Image.LANCZOS)
                img.save(output_path, dpi=(300, 300))
            print(f"Processed image {idx + 1}")

def manage_downloaded_folder(product_name, downloads_folder, base_output_folder, product_type):
    """
    Manages the downloaded folder:
    - Finds the latest ZIP file in the downloads folder.
    - Extracts it to a designated raw folder.
    - Creates a processed folder for processing images.
    - Moves the raw folder into the correct path.
    """
    try:
        # Step 1: Identify the latest ZIP file
        zip_file = max(pathlib.Path(downloads_folder).glob("*.zip"), key=os.path.getctime)
        print(f"Found downloaded ZIP file: {zip_file}")

        # Step 2: Define raw and processed folder paths
        raw_folder_name = f"raw {sanitize_name(product_name)}"
        raw_folder_path = os.path.join(base_output_folder, raw_folder_name)
        processed_folder_name = f"{sanitize_name(product_name)}-DPS"
        processed_folder_path = os.path.join(
            SEAMLESS_PATTERN_FOLDER if product_type == "Seamless Pattern" else DIGITAL_PAPER_FOLDER,
            processed_folder_name
        )

        # Step 3: Extract ZIP to raw folder
        if not os.path.exists(raw_folder_path):
            os.makedirs(raw_folder_path)
        shutil.unpack_archive(str(zip_file), raw_folder_path)
        os.remove(zip_file)  # Cleanup ZIP file
        print(f"✅ Extracted ZIP to raw folder: {raw_folder_path}")

        # Step 4: Ensure processed folder exists
        if not os.path.exists(processed_folder_path):
            os.makedirs(processed_folder_path)
        print(f"✅ Created processed folder: {processed_folder_path}")

        return raw_folder_path, processed_folder_path
    except Exception as e:
        logging.error(f"Error managing downloaded folder: {e}")

# ==============================
# Workflow for Single Product
# ==============================

def process_product(driver, product_data, product_index):
    """
    Full workflow for a single product:
    - Submits prompts to MidJourney.
    - Waits for images, downloads and processes them.
    """
    try:
        product_name = f"product {product_index+1}"
        print(f"🚀 Starting processing for: {product_name}")

        # Submit prompts to MidJourney
        send_prompts_to_midjourney(driver, [product_data])



        print(f"✅ Product {product_name} processed successfully!")
    except Exception as e:
        logging.error(f"Error processing product {product_name}: {e}")
        raise

# ==============================
# Main Workflow for All Products
# ==============================

def process_all_products():
    data = read_prompts_from_excel(INPUT_EXCEL_FILE)
    if not data:
        print("No valid data found in the Excel sheet.")
        return

    driver = connect_to_existing_edge()
    for idx, product_data in enumerate(data):
        try:
            process_product(driver, product_data, idx)
        except Exception as e:
            logging.error(f"⚠️ Process halted for product {idx+1}: {e}")
            backup_path = os.path.join(
                r"C:\Users\ULTRAPC\Desktop\Digital Paper Store\Digital Paper Store - Excel Sheet - Backup",
                "progress_backup.xlsx"
            )
            workbook = openpyxl.load_workbook(INPUT_EXCEL_FILE)
            workbook.save(backup_path)
            print(f"⚠️ Backup saved to {backup_path}")
            break
    print("✅ All products completed.")

# ==============================
# Main Execution
# ==============================
if __name__ == "__main__":
    process_all_products()