import os
import openpyxl
import time
from services.midjourney.process import update_excel_with_results

def test_excel_update():
    EXCEL_FILE = "template (4).xlsx"
    
    # Create unique product name using timestamp
    timestamp = int(time.time())
    unique_product_name = f"Test Product {timestamp}"
    
    # Test data with all fields that match the Excel columns
    test_product = {
        'Product Name': unique_product_name,  # Use unique name
        'Product Type': 'Seamless Pattern',
        'Category': 'Test Category',
        'Theme': 'Test Theme',
        'Prompts': [
            'A beautiful seamless pattern with flowers',
            'Vibrant colors, detailed petals',
            'Professional design, high quality'
        ],
        'Title': 'Beautiful Floral Seamless Pattern',
        'Hook': 'Transform your designs with this stunning floral pattern!',
        'Premade Description': 'A professionally designed seamless pattern featuring beautiful flowers.',
        'Full Description': '''This high-quality seamless pattern features:
- Stunning floral designs
- Vibrant, eye-catching colors
- Perfect for fabric, wallpaper, and more
- Professional quality artwork
- Instant digital download'''
    }
    
    # Test paths that match the expected structure
    raw_path = f"outputs/raw/test_product_{timestamp}"  # Use timestamp in paths too
    processed_path = f"outputs/processed/test_product_{timestamp}"
    share_link = f"https://drive.google.com/drive/folders/test-folder-id-{timestamp}?usp=sharing"
    
    print("\n🔍 Testing Excel Update")
    print("======================")
    
    # First, add a test row to the Excel file
    print("Adding test row to Excel...")
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook.active
    
    # Find the first empty row
    row = 2  # Start from row 2 (assuming row 1 is header)
    while sheet.cell(row=row, column=1).value:  # Assuming Product Name is in column A
        row += 1
    
    # Add test data
    sheet.cell(row=row, column=1, value=test_product['Product Name'])
    workbook.save(EXCEL_FILE)
    print(f"Added test row at row {row}")
    
    # Print test data for verification
    print("\nTest Data:")
    print("-----------")
    print(f"Product Name: {test_product['Product Name']}")
    print(f"Category: {test_product['Category']}")
    print(f"Theme: {test_product['Theme']}")
    print("Prompts:")
    for prompt in test_product['Prompts']:
        print(f"  - {prompt}")
    print(f"Title: {test_product['Title']}")
    print(f"Hook: {test_product['Hook']}")
    print("Premade Description:", test_product['Premade Description'])
    print("Full Description:", test_product['Full Description'])
    print("\nPaths:")
    print("------")
    print(f"Raw Path: {raw_path}")
    print(f"Processed Path: {processed_path}")
    print(f"Share Link: {share_link}")
    
    try:
        # Attempt to update the Excel file
        update_excel_with_results(test_product, raw_path, processed_path, share_link)
        
        print("\n✅ Excel update test completed!")
        print(f"Please check row {row} in template (4).xlsx to verify these updates:")
        print("1. All product information is present")
        print("2. Prompts are properly formatted")
        print("3. Paths are correctly set")
        print("4. Share link is valid")
        print("\nColumn mapping should be:")
        print("A: Product Name")
        print("B: Product Type")
        print("C: Category")
        print("D: Theme")
        print("E: Prompts")
        print("F: Raw Folder Path")
        print("G: Processed Folder Path")
        print("H: Google Drive Link")
        print("I: Listing Images Folder")
        print("J: Downloadable File Path")
        print("K: Title")
        print("L: Hook")
        print("M: Premade Description")
        print("N: Full Description")
        
    except Exception as e:
        print(f"\n❌ Test failed with error: {e}")

if __name__ == "__main__":
    test_excel_update() 